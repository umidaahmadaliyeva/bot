import os
import logging
import datetime
from telegram import (
    Update,
    ReplyKeyboardRemove,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.ext import (
    Application,
    ApplicationBuilder,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters,
)
from openpyxl import Workbook, load_workbook

# ================== SOZLAMALAR ==================
TOKEN = os.getenv("BOT_TOKEN")
CHANNEL_USERNAME = "@kh_journey"   # majburiy obuna kanali
EXCEL_FILE = "data.xlsx"

if not TOKEN:
    raise ValueError("BOT_TOKEN environment o'zgaruvchisi topilmadi! Railway Variables bo'limiga qo'shing.")

# ================== LOGGING ==================
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ================== HOLATLAR ==================
SCHOOL, CLASS_GRADE, FULL_NAME = range(3)

# ================== KANALGA OBUNA TEKSHIRISH ==================
async def check_subscription(user_id: int, bot) -> bool:
    try:
        member = await bot.get_chat_member(CHANNEL_USERNAME, user_id)
        return member.status in ["member", "administrator", "creator"]
    except Exception as e:
        logger.error(f"Obuna tekshirish xatosi: {e}")
        return False

# ================== EXCELGA SAQLASH ==================
def save_to_excel(data: dict):
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Vaqt", "Telegram ID", "Username",
                "Maktab", "Sinf", "Ism Familiya"
            ])
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        ws.append([
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            data["telegram_id"],
            data.get("username", "yo'q"),
            data["school"],
            data["class_grade"],
            data["full_name"],
        ])

        wb.save(EXCEL_FILE)
        logger.info(f"Ma'lumot saqlandi: {data['full_name']}")
    except Exception as e:
        logger.error(f"Excel saqlash xatosi: {e}")

# ================== /START ==================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    user_id = user.id

    is_subscribed = await check_subscription(user_id, context.bot)

    if not is_subscribed:
        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton(
                    text="📢 Kanalga obuna bo‘lish",
                    url=f"https://t.me/{CHANNEL_USERNAME.lstrip('@')}"
                )
            ],
            [
                InlineKeyboardButton(
                    text="✅ Tekshirish",
                    callback_data="check_sub"
                )
            ]
        ])

        await update.message.reply_text(
            "❗ Botdan foydalanish uchun avval kanalga obuna bo‘ling:",
            reply_markup=keyboard
        )
        return ConversationHandler.END

    # Tozalash va boshlash
    context.user_data.clear()
    context.user_data.update({
        "telegram_id": user_id,
        "username": user.username,
    })

    await update.message.reply_text(
        "Siz qaysi maktab o‘quvchisisiz?",
        reply_markup=ReplyKeyboardRemove(),
    )
    return SCHOOL

# ================== CALLBACK: OBUNANI TEKSHIRISH ==================
async def check_sub_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user = query.from_user
    is_subscribed = await check_subscription(user.id, context.bot)

    if not is_subscribed:
        await query.message.reply_text(
            "❌ Siz hali kanalga obuna bo‘lmagansiz.\n"
            "Iltimos, obuna bo‘lib yana tekshiring."
        )
        return

    context.user_data.clear()
    context.user_data.update({
        "telegram_id": user.id,
        "username": user.username,
    })

    await query.message.reply_text(
        "✅ Obuna muvaffaqiyatli tasdiqlandi!\n\n"
        "Siz qaysi maktab o‘quvchisisiz?"
    )

    # Suhbatni davom ettirish uchun holatni o'rnatish shart emas, chunki yangi start kabi
    return SCHOOL   # agar conversation davom etishi kerak bo'lsa

# ================== MAKTAB ==================
async def receive_school(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["school"] = update.message.text.strip()
    await update.message.reply_text("Siz nechinchi sinf o‘quvchisisiz?")
    return CLASS_GRADE

# ================== SINF ==================
async def receive_class(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["class_grade"] = update.message.text.strip()
    await update.message.reply_text("Ism va familiyangizni kiriting:")
    return FULL_NAME

# ================== ISM FAMILIYA ==================
async def receive_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["full_name"] = update.message.text.strip()

    save_to_excel(context.user_data)

    await update.message.reply_text(
        "✅ Sizning ma’lumotlaringiz saqlandi.\n"
        "Olimpiadada omad tilaymiz!"
    )
    return ConversationHandler.END

# ================== BEKOR QILISH ==================
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Ro‘yxatdan o‘tish bekor qilindi.")
    return ConversationHandler.END

# ================== ASOSIY QISM ==================
def main():
    logger.info("Bot ishga tushmoqda...")

    application = (
        ApplicationBuilder()
        .token(TOKEN)
        .connect_timeout(30)
        .read_timeout(30)
        .write_timeout(30)
        .get_updates_connect_timeout(30)
        .build()
    )

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            SCHOOL: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_school)],
            CLASS_GRADE: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_class)],
            FULL_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_name)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,  # qayta /start bosilsa davom etishi mumkin
    )

    application.add_handler(conv_handler)
    application.add_handler(CallbackQueryHandler(check_sub_callback, pattern="^check_sub$"))

    logger.info("Polling boshlanmoqda...")
    application.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=True,
        poll_interval=0.5,
        timeout=20,
        bootstrap_retries=-1,   # cheksiz qayta urinish
    )


if __name__ == "__main__":
    try:
        print("BOT_TOKEN mavjudmi?", bool(TOKEN))
        print("Polling boshlanmoqda...")
        main()
    except Exception as e:
        logger.error("Kritik xato:", exc_info=True)
        print("KRITIK XATO:", str(e))
        raise